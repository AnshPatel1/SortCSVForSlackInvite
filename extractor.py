import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

comment = '''
# 0.     Date submitted
# 1.     Identification
# 2.     ARE YOU A STUDENT OF SAL INSTITUTE OF DIPLOMA STUDIES?
# 3.     WHAT'S YOUR FIELD?
# 4.     WHAT'S YOUR CURRENT SEMESTER?
# 5.     WHAT'S YOUR FULL NAME?
# 6.     WHAT'S YOUR EMAIL?
# 7.     WHAT'S YOUR CONTACT NUMBER?
# 8.     WHAT'S YOUR ENROLLMENT NO?
# 9.     YOUR GitHub ACCOUNT LINK


# sr.no,    name,   enrollment,     number,     semester,    field,     email,      github
# auto,     5,      8,              7,          4,           3,         6,          9

# list = [[auto, 5, 8, 7, 4, 3, 6, 9], [], []]
'''


def extract_emails_as_list(filename):
    cursor = 1
    isFirstRowSkipped = False
    final_list = [
        ["Sr. No.", "Name", "Enrollment No.", "Mobile Number", "Semester", "Field", "Email", "GitHub Account Link"]]
    with open(filename) as f:
        csv_file = csv.reader(f)
        for row in csv_file:
            if isFirstRowSkipped is False:
                isFirstRowSkipped = True
                continue
            if isFirstRowSkipped:
                final_list.append(
                    [str(cursor), str(row[5]).upper(), str(row[8]), str(row[7]), str(row[4]), str(row[3]), str(row[6]),
                     str(row[9])])
                cursor = cursor + 1
    return final_list


def generate_excel(pre_sorted_list):
    wb = Workbook()
    dest_name = "Result.xlsx"
    #     create a new worksheet
    sheet1 = wb.active
    isFirstRowSkipped = False
    for line in pre_sorted_list:
        row_number = 1
        if isFirstRowSkipped is False:
            isFirstRowSkipped = True
            line = ['Sr. No.', 'Name', 'Enrollment No.', 'Mobile Number', 'Semester', 'Field', 'Email',
                    'GitHub Account Link']
        else:
            row_number = int(line[0]) + 1

        letter_no = 0
        for item in line:
            letter_no = letter_no + 1
            letter = get_column_letter(letter_no)
            sheet1[letter + str(row_number)] = str(item)

    wb.save(dest_name)


def find_students_without_github(pre_sorted_list):
    no_github_list = []
    for row in pre_sorted_list:
        if row[7] == '':
            no_github_list.append(row[1])
    return no_github_list


def generate_invite_list(lists):
    invite_list = ''
    for row in lists:
        if row[6] != 'Email':
            invite_list = invite_list + row[6] + ', '
    return invite_list


def store_mail_list_to_file(email_list):
    with open('invite_list.txt', 'w') as list_file:
        list_file.write(email_list)


if __name__ == "__main__":
    sorted_list = extract_emails_as_list('form.csv')
    # generate_excel(sorted_list)
    # for n in find_students_without_github(sorted_list):
    #     print(n)
    store_mail_list_to_file(generate_invite_list(sorted_list))
